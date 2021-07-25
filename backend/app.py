from flask import Flask, request, send_file
from flask_cors import CORS
import openpyxl as xl
from japanera import EraDate

# flask
app = Flask(__name__)
CORS(app)

# Office system ファイル形式のMIMEタイプをサーバに登録
XLSX_MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

# rest api
# json with base64 encoding
# {
#    'fileName': 'ファイル名',
#    'contentType': 'mimetype',
#    'contentData': 'バイナリデータをbase64でエンコードしたascii文字',
# }
@app.route('/download', methods=['POST'])
def create_PDF():
    data = request.get_json()
    downloadFileName = 'Deduction_application.xlsx'
    downloadFile = 'Deduction_application.xlsx'
    createExcel(downloadFileName, data['basicInfo'], data['incomeCal'], data['incomeAdjust'])
    return send_file(downloadFile, as_attachment=True,
        attachment_filename=downloadFileName,
        mimetype=XLSX_MIMETYPE)

def Classification1(incomeQuote):
    if incomeQuote > 25000000:
        return '', ''
    elif incomeQuote > 24500000:
        return '', '16 万円'
    elif incomeQuote > 24000000:
        return '', '32 万円'
    elif incomeQuote > 10000000:
        return '', '48 万円'
    elif incomeQuote > 9500000:
        return 'C', '48 万円'
    elif incomeQuote > 9000000:
        return 'B', '48 万円'
    return 'A', '48 万円'

def Classification2(incomeQuote, year):
    if incomeQuote <= 480000 and year <= 1951:
        return '①'
    elif incomeQuote <= 480000 and year > 1951:
        return '②'
    elif incomeQuote <= 950000:
        return '③'
    elif incomeQuote <= 1330000:
        return '④'
    return ''

def PartnerDeduction(classification1, classification2, incomeQuote):
    if classification2 == '①':
        if classification1 == 'A':
            return 0, '48万円'
        elif classification1 == 'B':
            return 0, '32万円'
        elif classification1 == 'C':
            return 0, '16万円'
    elif classification2 == '②':
        if classification1 == 'A':
            return 0, '38万円'
        elif classification1 == 'B':
            return 0, '26万円'
        elif classification1 == 'C':
            return 0, '13万円'
    elif classification2 == '③':
        if classification1 == 'A':
            return 1, '38万円'
        elif classification1 == 'B':
            return 1, '26万円'
        elif classification1 == 'C':
            return 1, '13万円'
    elif classification2 == '④':
        if incomeQuote > 900000:
            if classification1 == 'A':
                return 1, '38万円'
            elif classification1 == 'B':
                return 1, '24万円'
            elif classification1 == 'C':
                return 1, '12万円'
        elif incomeQuote > 1000000:
            if classification1 == 'A':
                return 1, '31万円'
            elif classification1 == 'B':
                return 1, '21万円'
            elif classification1 == 'C':
                return 1, '11万円'
        elif incomeQuote > 1050000:
            if classification1 == 'A':
                return 1, '26万円'
            elif classification1 == 'B':
                return 1, '18万円'
            elif classification1 == 'C':
                return 1, '9万円'
        elif incomeQuote > 1100000:
            if classification1 == 'A':
                return 1, '21万円'
            elif classification1 == 'B':
                return 1, '14万円'
            elif classification1 == 'C':
                return 1, '7万円'
        elif incomeQuote > 1150000:
            if classification1 == 'A':
                return 1, '16万円'
            elif classification1 == 'B':
                return 1, '11万円'
            elif classification1 == 'C':
                return 1, '6万円'
        elif incomeQuote > 1200000:
            if classification1 == 'A':
                return 1, '11万円'
            elif classification1 == 'B':
                return 1, '8万円'
            elif classification1 == 'C':
                return 1, '4万円'
        elif incomeQuote > 1250000:
            if classification1 == 'A':
                return 1, '6万円'
            elif classification1 == 'B':
                return 1, '4万円'
            elif classification1 == 'C':
                return 1, '2万円'
        elif incomeQuote > 1300000:
            if classification1 == 'A':
                return 1, '3万円'
            elif classification1 == 'B':
                return 1, '2万円'
            elif classification1 == 'C':
                return 1, '1万円'
        elif incomeQuote >= 1330000:
            return -1, ''
    return -1, ''


def createExcel(fileName, basicInfo, incomeCal, incomeAdjust):
    wb = xl.load_workbook('template.xlsx')
    ws = wb['申告書']
    # 基本情報
    ws['R10'] = basicInfo["school"]
    ws['R8'] = basicInfo["stuffNum"]
    ws['AP6'] = basicInfo["stuffLNRuby"] + "　" + basicInfo["stuffFNRuby"]
    ws['AP7'] = basicInfo["stuffLN"] + "　" + basicInfo["stuffFN"]
    ws['AP10'] = basicInfo["stuffAdr"]
    ws['AX24'] = basicInfo["partnerNum"]
    ws['AX24'] = basicInfo["partnerNum"]
    ws['AI27'] = basicInfo["partnerLNRuby"] + "　" + basicInfo["partnerFNRuby"]
    ws['AI29'] = basicInfo["partnerLN"] + "　" + basicInfo["partnerFN"]
    # 生年月日
    if basicInfo["partnerBD"] != '':
        bd = basicInfo["partnerBD"].split('/')
        era_date = EraDate(int(bd[0]), int(bd[1]), int(bd[2]))
        ws['BM24'] = era_date.strftime("%-E%-O年%m月%d日")

    ws['AX29'] = basicInfo["partnerAdr"]
    if basicInfo["partnerResidnet"] == True:
        ws['BM29'] = "〇"
    ws['BS29'] = basicInfo["partnerEvid"]

    # 所得金額計算
    ws['L38'] = incomeCal["earnings1"]
    ws['V38'] = incomeCal["income1"]
    ws['V43'] = incomeCal["exIncome1"]
    incomeSum1 = incomeCal["income1"] + incomeCal["exIncome1"]
    ws['V48'] = incomeSum1
    classification1, deduction1 = Classification1(incomeSum1)
    ws['Y56'] = classification1
    ws['Y63'] = deduction1
    # 配偶者
    ws['AQ38'] = incomeCal["earnings2"]
    ws['BA38'] = incomeCal["income2"]
    ws['BA43'] = incomeCal["exIncome2"]
    incomeSum2 = incomeCal["income2"] + incomeCal["exIncome2"]
    ws['BA48'] = incomeSum2
    if incomeCal["income2"] != 0 and basicInfo["partnerBD"] != '':
        bd = basicInfo["partnerBD"].split('/')
        classification2 = Classification2(incomeSum2, int(bd[0]))
        n, deduction2 = PartnerDeduction(classification1, classification2, incomeSum2)
        if n == 0:
            ws['BU56'] = deduction2
        elif n == 1:
            ws['BU63'] = deduction2

    # 所得金額調整
    if incomeAdjust["radioGroup"] == "2":
        ws['F82'].value = "☑" + ws['F82'].value[1:]
    elif incomeAdjust["radioGroup"] == "3":
        ws['F84'].value = "☑" + ws['F84'].value[1:]
    elif incomeAdjust["radioGroup"] == "4":
        ws['F86'].value = "☑" + ws['F86'].value[1:]
    elif incomeAdjust["radioGroup"] == "5":
        ws['F88'].value = "☑" + ws['F88'].value[1:]
    
    if incomeAdjust["dependentsNum"] != 0:
        ws['AQ83'] = incomeAdjust["dependentsNum"]
    if incomeAdjust["dependentsBD"] != '':
        bd = incomeAdjust["dependentsBD"].split('/')
        era_date = EraDate(int(bd[0]), int(bd[1]), int(bd[2]))
        ws['BC83'] = era_date.strftime("%-E%-O年%m月%d日")
    ws['AE86'] = incomeAdjust["dependentsLNR"] + "　" + incomeAdjust["dependentsFNR"]
    ws['AE88'] = incomeAdjust["dependentsLN"] + "　" + incomeAdjust["dependentsFN"]
    ws['AQ88'] = incomeAdjust["dependentsAdr"]
    ws['BC88'] = incomeAdjust["dependentsRel"]
    ws['BH88'] = incomeAdjust["dependentsInc"]
    ws['BR84'] = incomeAdjust["disabilityPrsEvid"]
    wb.save(fileName)

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=8000, debug=True)