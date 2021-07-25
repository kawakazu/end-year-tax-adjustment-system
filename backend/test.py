import openpyxl as xl
from datetime import date
from japanera import Japanera, EraDate

basicInfo = {
    "school": "西小学校",
    "stuffNum": 123456789,
    "stuffLN": "川口",
    "stuffFN": "和久",
    "stuffLNRuby": "カワグチ",
    "stuffFNRuby": "カズヒサ",
    "stuffAdr": "三重県名張市桔梗が丘", 
    "partnerNum": 123456789,
    "partnerLN": "川口",
    "partnerFN": "和久",
    "partnerLNRuby": "カワグチ",
    "partnerFNRuby": "カズヒサ",
    "partnerAdr": "三重県名張市桔梗が丘",
    "partnerBD": "1996/10/07",
    "partnerResidnet": "0",
    "partnerEvid": "普通免許",
}

incomCal = {
    "earnings1": 8750000,
    "income1": 6800000,
    "exIncome1": 275000,
    # 給与所得の必要経費
    # 給与以外の合計
    "earnings2": 1100000,
    "income2": 550000,
    "exIncome2": 0
}

incomAdjust = {
    "radioGroup": "2",
    "dependentsNum": 1234565789,
    "dependentsBD": "1996/10/07",
    "dependentsLN": "川口",
    "dependentsFN": "和久",
    "dependentsLNR": "カワグチ",
    "dependentsFNR": "カズヒサ",
    "dependentsAdr": "三重県名張市桔梗が丘",
    "dependentsRel": "子",
    "dependentsInc": 100000,
    "disabilityPrsEvid": "普通免許"
}

#def ExchangeJapaneseCalendar(year):


def Classification1(incomeQuote):
    if incomeQuote > 25000000:
        return '', ''
    elif incomeQuote > 24500000:
        return '', '16 万円'
    elif incomeQuote > 24000000:
        return '', '32 万円'
    elif incomeQuote > 10000000:
        return '', '48 万円'
    elif incomeQuote > 9500000:
        return 'C', '48 万円'
    elif incomeQuote > 9000000:
        return 'B', '48 万円'
    return 'A', '48 万円'

def Classification2(incomeQuote, year):
    if incomeQuote <= 480000 and year <= 1951:
        return '①'
    elif incomeQuote <= 480000 and year > 1951:
        return '②'
    elif incomeQuote <= 950000:
        return '③'
    elif incomeQuote <= 1330000:
        return '④'
    return ''


def createExcel():
    wb = xl.load_workbook('template.xlsx')
    ws = wb['申告書']
    # 基本情報
    ws['R10'] = basicInfo["school"]
    ws['R8'] = basicInfo["stuffNum"]
    ws['AP6'] = basicInfo["stuffLNRuby"] + "　" + basicInfo["stuffFNRuby"]
    ws['AP7'] = basicInfo["stuffLN"] + "　" + basicInfo["stuffFN"]
    ws['AP10'] = basicInfo["stuffAdr"]
    ws['AX24'] = basicInfo["partnerNum"]
    ws['AX24'] = basicInfo["partnerNum"]
    ws['AI27'] = basicInfo["partnerLNRuby"] + "　" + basicInfo["partnerFNRuby"]
    ws['AI29'] = basicInfo["partnerLN"] + "　" + basicInfo["partnerFN"]
    # 生年月日
    bd = basicInfo["partnerBD"].split('/')
    era_date = EraDate(int(bd[0]), int(bd[1]), int(bd[2]))
    ws['BM24'] = era_date.strftime("%-E%-O年%m月%d日")

    ws['AX29'] = basicInfo["partnerAdr"]
    if basicInfo["partnerResidnet"] == "1":
        ws['BM29'] = "〇"
    ws['BS29'] = basicInfo["partnerEvid"]

    # 所得金額計算
    ws['L38'] = incomCal["earnings1"]
    ws['V38'] = incomCal["income1"]
    ws['V43'] = incomCal["exIncome1"]
    incomeSum = incomCal["income1"] + incomCal["exIncome1"]
    ws['V48'] = incomeSum
    classification, deduction = Classification1(incomeSum)
    ws['Y56'] = classification
    ws['Y63'] = deduction
    # 配偶者
    ws['AQ38'] = incomCal["earnings2"]
    ws['BA38'] = incomCal["income2"]
    ws['BA43'] = incomCal["exIncome2"]
    ws['BA48'] = incomCal["income2"] + incomCal["exIncome2"]
    


    # 所得金額調整
    if incomAdjust["radioGroup"] == "2":
        ws['F82'].value = "☑" + ws['F82'].value[1:]
    elif incomAdjust["radioGroup"] == "3":
        ws['F84'].value = "☑" + ws['F84'].value[1:]
    elif incomAdjust["radioGroup"] == "4":
        ws['F86'].value = "☑" + ws['F86'].value[1:]
    elif incomAdjust["radioGroup"] == "5":
        ws['F88'].value = "☑" + ws['F88'].value[1:]
    ws['AQ83'] = incomAdjust["dependentsNum"]
    bd = incomAdjust["dependentsBD"].split('/')
    era_date = EraDate(int(bd[0]), int(bd[1]), int(bd[2]))
    ws['BC83'] = era_date.strftime("%-E%-O年%m月%d日")
    ws['AE86'] = incomAdjust["dependentsLNR"] + "　" + incomAdjust["dependentsFNR"]
    ws['AE88'] = incomAdjust["dependentsLN"] + "　" + incomAdjust["dependentsFN"]
    ws['AQ88'] = incomAdjust["dependentsAdr"]
    ws['BC88'] = incomAdjust["dependentsRel"]
    ws['BH88'] = incomAdjust["dependentsInc"]
    ws['BR84'] = incomAdjust["disabilityPrsEvid"]
    wb.save('test.xlsx')

createExcel()
